import os
import datetime
import openpyxl
from requests import *
from bs4 import *
from database import Database

def Download():
    url = 'https://kcpt72.ru/schedule/'
    urldownload = ParsKCPT(url)

    excelfile = []
    wordfile = ""
    for i in urldownload:
        filename = str(i).split('/')
        if filename[3] == 'spreadsheets':
            excelfile.append(f"https://drive.google.com/u/0/uc?id={filename[5]}&export=download")

    path = "Excel"

    num = 1
    for i in excelfile:
        response = get(i)
        # сохраните файл в папке с заданным именем
        with open(os.path.join(path, f'Schedule{num}.xlsx'), 'wb') as f:
            f.write(response.content)
        num += 1
    
    twotable("Excel/Schedule2.xlsx", "Excel/Schedule1.xlsx")

def ParsKCPT(url):

    arr_a_google = []
    response = get(url, verify=False)
    html_content = response.text


    soup = BeautifulSoup(html_content, 'html.parser')
    a_tags = soup.find_all("a")

    for i in a_tags:
        if "docs.google.com" in str(i):
            ps = BeautifulSoup(str(i), "html.parser")
            link = ps.find('a')['href']
            arr_a_google.append(link)
    
    return arr_a_google

def twotable(pathfile1: str, pathfile2: str):
    # Читаем первый файл
    workbook1 = openpyxl.load_workbook(pathfile1, data_only=True)

    # Читаем второй файл
    workbook2 = openpyxl.load_workbook(pathfile2, data_only=True)

    # Создаем новый Workbook
    merged_workbook = openpyxl.Workbook()
    merged_sheet = merged_workbook.active

    # Сливаем данные из первого файла
    for sheet_name in workbook1.sheetnames:
        sheet = workbook1[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            merged_sheet.append(row)

    # Сливаем данные из второго файла
    for sheet_name in workbook2.sheetnames:
        sheet = workbook2[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            merged_sheet.append(row)

    # Сохраняем объединенный Workbook в новый файл
    merged_workbook.save("Excel/2table.xlsx")

def AddToDatabase(Raspisanie:dict):
    for date,rasp in Raspisanie.items():
        for group,urok in rasp.items():
            #Вот отсюда можно добавлять группы (print(group))
            for i in urok:
                while None in i:
                    i.remove(None)
                if len(i) == 4:
                    Database.AddUrok(
                        Number=i[0],
                        Subject=i[1],
                        Classroom=i[2],
                        Prepod=i[3],
                        Date=datetime.datetime.strptime(date, '%d.%m.%Y').date(),
                        Group=group
                    )
                elif len(i) == 3:
                    Database.AddUrok(
                        Number=i[0],
                        Subject=i[1],
                        Classroom='-',
                        Prepod=i[2],
                        Date=datetime.datetime.strptime(date, '%d.%m.%Y').date(),
                        Group=group
                    )
                elif len(i) == 6:
                    Database.AddUrok(
                        Number=i[0],
                        Subject=i[1],
                        Classroom=i[2],
                        Prepod=i[3],
                        Date=datetime.datetime.strptime(date, '%d.%m.%Y').date(),
                        Group=group
                    )
                    Database.AddUrok(
                        Number=i[0],
                        Subject=i[4],
                        Classroom='',
                        Prepod=i[5],
                        Date=datetime.datetime.strptime(date, '%d.%m.%Y').date(),
                        Group=group
                    )
                    
                elif len(i) == 7:
                    Database.AddUrok(
                        Number=i[0],
                        Subject=i[1],
                        Classroom=i[2],
                        Prepod=i[3],
                        Date=datetime.datetime.strptime(date, '%d.%m.%Y').date(),
                        Group=group
                    )
                    Database.AddUrok(
                        Number=i[0],
                        Subject=i[4],
                        Classroom=i[5],
                        Prepod=i[6],
                        Date=datetime.datetime.strptime(date, '%d.%m.%Y').date(),
                        Group=group
                    )
                elif len(i) == 9:
                    Database.AddUrok(
                        Number=i[0],
                        Subject=i[1],
                        Classroom=i[2],
                        Prepod=i[3],
                        Date=datetime.datetime.strptime(date, '%d.%m.%Y').date(),
                        Group=group
                    )
                    Database.AddUrok(
                        Number=i[0],
                        Subject=i[4],
                        Classroom='-',
                        Prepod=i[5],
                        Date=datetime.datetime.strptime(date, '%d.%m.%Y').date(),
                        Group=group
                    )
                    Database.AddUrok(
                        Number=i[0],
                        Subject=i[6],
                        Classroom=i[7],
                        Prepod=i[8],
                        Date=datetime.datetime.strptime(date, '%d.%m.%Y').date(),
                        Group=group
                    )
                elif len(i) == 10:
                    Database.AddUrok(
                        Number=i[0],
                        Subject=i[1],
                        Classroom=i[2],
                        Prepod=i[3],
                        Date=datetime.datetime.strptime(date, '%d.%m.%Y').date(),
                        Group=group
                    )
                    Database.AddUrok(
                        Number=i[0],
                        Subject=i[4],
                        Classroom=i[5],
                        Prepod=i[6],
                        Date=datetime.datetime.strptime(date, '%d.%m.%Y').date(),
                        Group=group
                    )
                    Database.AddUrok(
                        Number=i[0],
                        Subject=i[7],
                        Classroom=i[8],
                        Prepod=i[9],
                        Date=datetime.datetime.strptime(date, '%d.%m.%Y').date(),
                        Group=group
                    )

def ParsTwoTable():
    workbook = openpyxl.load_workbook('Excel/2table.xlsx')
    sheet = workbook.active
    unformated = []
    raspis = []
    for column in sheet.iter_rows(values_only=True):
        if column[0] != None and type(column[0]) != int and "День" in column[0]:
            if raspis != []:
                unformated.append(raspis)
            raspis = []
            raspis.append(column[0].split(",")[1])
        if column[0] == "№":
            for i in column:
                if i != "№" and i != None:
                    raspis.append([i])
        if type(column[0]) == int:
            listColumn = list(column)
            for i in raspis[1:]:
                i.append([listColumn[0]])
                i[-1].append(listColumn[1])
                i[-1].append(listColumn[2])
                del listColumn[1]
                del listColumn[1]
        if column[0] == None and column[2] != "Ауд." :
            listDop = list(column)
            for i in raspis[1:]:
                i[-1].append(listDop[1])
                i[-1].append(listDop[2])
                del listDop[1]
                del listDop[1]

    output = {}
    for i in unformated:
        output[i[0].strip()] = []
    for i in unformated:
        output[i.pop(0).strip()].append(i)
    for date in output.items():
        rasdate = {}
        raspis = date[1]
        for i in raspis:
            for j in i:
                rasdate[j.pop(0)] = j
        output[date[0]] = rasdate

    os.remove("Excel/Schedule1.xlsx")
    os.remove("Excel/Schedule2.xlsx")
    os.remove("Excel/2table.xlsx")
    return output
#Download()
#AddToDatabase(ParsTwoTable())
#Database.RemoveOldSchedule()