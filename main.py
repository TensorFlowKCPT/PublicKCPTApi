from sanic import Sanic, response, json, html
from sanic import Sanic
from sanic.response import html
from jinja2 import Environment, FileSystemLoader, select_autoescape
from database import Database
from ParsKCPT import Download, AddToDatabase, ParsTwoTable
import threading
import datetime
import re
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font

app = Sanic("KCPTApi")

link = "https://localhost:8000"

env = Environment(
    loader=FileSystemLoader('templates'),  # Папка с шаблонами
    autoescape=select_autoescape(['html', 'xml'])
)
# Создаем словарь для преобразования чисел месяцев в текстовое представление на русском языке
months = {
    1: 'января',
    2: 'февраля',
    3: 'марта',
    4: 'апреля',
    5: 'мая',
    6: 'июня',
    7: 'июля',
    8: 'августа',
    9: 'сентября',
    10: 'октября',
    11: 'ноября',
    12: 'декабря'
}

# Создаем словарь для преобразования дней недели в текстовое представление на русском языке
days_of_week = {
    0: 'понедельник',
    1: 'вторник',
    2: 'среда',
    3: 'четверг',
    4: 'пятница',
    5: 'суббота',
    6: 'воскресенье'
}

@app.route('/api/changes_day')
async def apiChangesDay(request):
    return json(Database.GetChangesDay(datetime.datetime.strptime(request.args.get('date'),'%Y-%m-%d').date()))

@app.route('/api/class_day')
async def apiClassDay(request):
    return json(Database.GetScheduleWithChangesDay(datetime.datetime.strptime(request.args.get('date'),'%Y-%m-%d').date(),request.args.get('group')))

@app.route('/api/class_week')
async def apiClassWeek(request):
    output = {}
    date = datetime.datetime.strptime(request.args.get('date'),'%Y-%m-%d').date()
    group = request.args.get('group')
    for i in range(7):
        output[(date+ datetime.timedelta(days=i)).strftime('%Y-%m-%d')]=(Database.GetScheduleWithChangesDay(date+ datetime.timedelta(days=i), group))
    return json(output)

@app.route('/api/teacher_day')
async def apiTeacherDay(request):
        return json(Database.GetPrepodsDay(datetime.datetime.strptime(request.args.get('date'),'%Y-%m-%d').date(),request.args.get('teacher')))

@app.route('api/teacher_week')
async def apiTeacherWeek(request):
        output = {}
        date = datetime.datetime.strptime(request.args.get('date'),'%Y-%m-%d').date()
        for i in range(7):
            output[(date+ datetime.timedelta(days=i)).strftime('%Y-%m-%d')]=(Database.GetPrepodsDay((date+datetime.timedelta(days=i)),request.args.get('teacher')))
        return json(output)

@app.route('/api/classrooms')
async def apiClassroomsList(request):
    return json(Database.GetAllClassrooms())

@app.route('/api/subjects')
async def apiSubjectsList(request):
    return json(Database.GetAllSubjects())

@app.route('/api/teachers')
async def apiTeachersList(request):
    return json(Database.GetPrepodsList())

@app.route('/api/groups')
async def apiGroupsList(request):
    return json(Database.GetAllGroups())

@app.route('/api')
async def apiDocs(request):
    link = request.url
    Data = {'link': link, 'today': datetime.datetime.now().date().strftime('%Y-%m-%d')}
    template = env.get_template('ApiDocs.html')
    return response.html(template.render(data = Data))

@app.route("/")
async def index(request):
    Data = {}
    template = env.get_template('index.html')
    return response.html(template.render(data = Data))

@app.post('/admin/ParsKCPT')
async def StartParser(request):
    Download()
    AddToDatabase(ParsTwoTable())
    Database.RemoveOldSchedule()
    return response.redirect('/admin/ChangesPage')

def custom_enumerate(data):
    return enumerate(data)


@app.route('/admin/ChangesPage')
async def adminPanelChanges(request):
    date = request.args.get('date')
    if date != None:
        Changes = Database.GetChangesDay(datetime.datetime.strptime(date,'%Y-%m-%d').date())
        try:
            Changes = list(Changes.values())
        except:
            Data = {'date':date,'Changes': None, 'Groups' : Database.GetAllGroups(), 'Prepods' : sorted(Database.GetAllPrepods()), 'Subjects':Database.GetAllSubjects(), 'Classrooms' : Database.GetAllClassrooms()}
            template = env.get_template('changesPage.html')
            resp = response.html(template.render(data = Data))
            #resp.add_cookie('date',date)
            return resp
        
        for i in Changes:
            if i['Delete'] != None:
                i['Scheduled'] = Database.GetUrokByDateNumberAndGroup(datetime.datetime.strptime(date,'%Y-%m-%d').date(), i['Group'], i['Delete'])
        Data = {'date':date,'Changes': Changes, 'Groups' : Database.GetAllGroups(), 'Prepods' : sorted(Database.GetAllPrepods()), 'Subjects':Database.GetAllSubjects(), 'Classrooms' : Database.GetAllClassrooms()}
    else:
        Data = {'date':date,'Changes': None, 'Groups' : Database.GetAllGroups(), 'Prepods' : sorted(Database.GetAllPrepods()), 'Subjects':Database.GetAllSubjects(), 'Classrooms' : Database.GetAllClassrooms()}
    template = env.get_template('changesPage.html')
    resp = response.html(template.render(data = Data))
    #resp.add_cookie('date',date)
    return resp

#@app.route('/admin/printChange')
#async def adminPanelChanges(request):
#    date = request.args.get('date')
#    Changes = Database.GetChangesDay(datetime.datetime.strptime(date,'%Y-%m-%d').date())
#    Changes = sorted(Changes,key=lambda x: x['Group'])
#    
#    WordGeneration.GenerateWord(Changes)
#    return resp

@app.route('/image/<filename:str>')
async def serve_image(request, filename):
    return await response.file('images/'+filename)

def GetChangesDict(request):
    date = request.args.get("date", None)
    Date = datetime.datetime.strptime(date,'%Y-%m-%d')
    if date:
        Changes = Database.GetChangesDay(datetime.datetime.strptime(date,'%Y-%m-%d').date())
        try:
            Changes = sorted(list(Changes.values()), key=lambda x: x['Group'])
        except:
            return response.text("Изменений на эту дату нет")
        for i in Changes:
            if i['Delete'] != None:
                 i['Scheduled'] = Database.GetUrokByDateNumberAndGroup(datetime.datetime.strptime(date,'%Y-%m-%d').date(), i['Group'], i['Delete'])
        pass
    Changes = sorted(list(Changes), key=lambda x: (x['Group'], x['Urok']))
    # Simulate PHP-style database querying
    # Replace this with actual database queries
    
        
    changes_data = Changes
    output_data = []
    unique_entries = {}

    for change in changes_data:
        key = (change['Group'], change['Classroom'], change['Subject'], change['Comment'])
        if key in unique_entries:
            unique_entries[key]['Urok'] = f"{str(unique_entries[key]['Urok']).split('-')[0]}-{change['Urok']}"
        else:
            unique_entries[key] = change


    output_data = list(unique_entries.values())

    Group = ''
    for i in output_data:
        if i['Group'] != Group:
            Group = i['Group']
        else:
            i['Group'] = ''
        if i['Comment'] == None:
            i['Comment'] = ''
        if i['Classroom'] == None:
            i['Classroom'] = ''
        if i['Subject'] == None:
            i['Subject'] = ''

    for i in output_data:
        if i['Comment'].lower() == 'отмена':
            try:
                i['Scheduled']['SubjectId'] = i['Subject']
            except:
                pass
            i['Subject'] = ''
            try:
                i['Scheduled']['Prepod'] = i['Prepod']
            except:
                pass
            i['Prepod'] = ''
            i['Classroom'] = '-'
        if i['Comment'].lower() == 'замена кабинета':
            if type(i['Scheduled']) != str:
                i['Scheduled']['SubjectId'] = i['Subject']
                i['Subject'] = ''
                try:
                    i['Scheduled']['Prepod'] = i['Prepod']
                except TypeError:
                    pass
                i['Prepod'] = ''
            else:
                i['Subject'] = ''
                i['Prepod'] = ''
    for i in output_data:
        try:
            text = re.match('(\d\.)?МДК\.[0-9]+\.[0-9]+', i['Subject'])[0]
        except:
            text = None
        try:
            scheduledtext = re.match('(\d\.)?МДК\.[0-9]+\.[0-9]+', i['Scheduled']['SubjectId'])[0]
        except:
            scheduledtext = None
        if text != None:
            i['Subject'] = text
        if scheduledtext != None:
            i['Scheduled']['SubjectId'] = scheduledtext
    return output_data


@app.route("/admin/printChanges")
async def printChanges(request):

    # Simulate PHP-style date processing
    date = request.args.get("date", None)
    Date = datetime.datetime.strptime(date,'%Y-%m-%d')
    day = Date.day
    month = months[Date.month]
    year = Date.year
    day_of_week_str = days_of_week[Date.weekday()]
    
    # Формируем строку в нужном формате
    formatted_date = f"{day} {month} {year} года ({day_of_week_str})"
    output_data = GetChangesDict(request)
    
    template = env.get_template("printChanges.html")
    rendered_template = template.render(date=formatted_date, changes_data=output_data)
    return html(rendered_template)
    
@app.route('/admin/printExcel')
async def printExcel(request):
    # Simulate PHP-style date processing
    date = request.args.get("date", None)
    Date = datetime.datetime.strptime(date,'%Y-%m-%d')
    day = Date.day
    month = months[Date.month]
    year = Date.year
    day_of_week_str = days_of_week[Date.weekday()]
    formatted_date = f"{day} {month} {year} года ({day_of_week_str})"
    output_data = GetChangesDict(request)
    # Создаем новую рабочую книгу
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.merge_cells("A1:E1")
    ws['A1'] = 'ГАПОУ ТО «Колледж цифровых и педагогических технологий»'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells("A2:E2")
    ws['A2'] = 'ИЗМЕНЕНИЯ В РАСПИСАНИИ'
    ws['A2'].font = Font(size=16, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells("A3:E3")
    ws['A3'] = formatted_date
    ws['A3'].font = Font(size=16, bold=True)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    

    # Заголовок таблицы
    ws['A4'] = 'Группа'
    ws['B4'] = '№ урока'
    ws['C4'] = 'По расписанию'
    ws['D4'] = 'Изменения'
    ws['E4'] = 'Кабинет'

    # Данные из словаря
    data = output_data

    row = 5
    for item in data:
        if item['IsChange']:
            ws.cell(row=row, column=1, value=item['Group'])
            ws.cell(row=row, column=2, value=item['Urok'])
            try:
                ws.cell(row=row, column=3, value=item['Scheduled']['SubjectId'] + "\n" + item['Scheduled']['Prepod'])
            except TypeError:
                pass
            value = item["Comment"]
            if item["Subject"] != '':
                value += '\n' + item["Subject"]
            if item["Prepod"] != '':
                value += '\n' + item["Prepod"]
            ws.cell(row=row, column=4, value=value)
            ws.cell(row=row, column=5, value=item['Classroom'])

            # Форматирование ячеек
            ws.row_dimensions[row].height = 60  # Высота строки
            ws.column_dimensions['A'].width = 15  # Ширина столбца A
            ws.column_dimensions['B'].width = 10  # Ширина столбца B
            ws.column_dimensions['C'].width = 20  # Ширина столбца C
            ws.column_dimensions['D'].width = 30  # Ширина столбца D
            ws.column_dimensions['E'].width = 15  # Ширина столбца E

        # Выравнивание текста
        for cell in ws[row]:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        
        if item['Group'].strip() == '':
            for cell in ws[row]:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='none'), bottom=Side(style='none'))
        else:
            for cell in ws[row]:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='none'))

    
        row += 1
    
    for cell in ws[row]:
        cell.border = Border(left=Side(style='none'), right=Side(style='none'), top=Side(style='thin'), bottom=Side(style='none'))
    
    ws.cell(row = row + 2, column = 1, value="Заместитель директора___________________Л.Г. Чудная")
    ws.cell(row = row + 4, column = 1, value="Диспетчер___________________О.В. Беловодская")
    # Сохраняем Excel-файл
    wb.save('schedule_formatted.xlsx')
    
    return await response.file('schedule_formatted.xlsx', headers={"Content-Disposition": f"attachment; filename={'schedule_formatted.xlsx'}"})

@app.route('/admin/GetChangeData')
async def GetChangeData(request):
    id = request.args.get('change')
    change = Database.GetChangeById(id)

    # Вывод всех данных в консоль
    #print('selectGroup:', str(change['Group']))
    #print('startSelect:', str(change['Urok']))
    #print('endSelect:', str(change['Urok']))
    #print('selectChangePrepod:', str(change['Prepod']))
    #print('selectChangeSubject:', str(change['Subject']))
    #print('changeStatusSelect:', str(change['Comment']))
    #print('selectChangeClassroom:', str(change['Classroom']))

    # Возвращение данных в виде JSON-ответа
    return response.json({
        'selectGroup': str(change['Group']),
        'startSelect': str(change['Urok']),
        'endSelect': str(change['Urok']),
        'selectChangePrepod': str(change['Prepod']),
        'selectChangeSubject': str(change['Subject']),
        'changeStatusSelect': str(change['Comment']),
        'selectChangeClassroom': str(change['Classroom'])
    })

@app.route('/admin/ShowSchedule')
async def adminShowSchedule(request):
    output = {}
    for i in Database.GetAllGroups():
        output[i] = (Database.GetScheduleWithChangesDay(datetime.datetime.strptime(request.args.get('date'), '%Y-%m-%d').date(), i))
    template = env.get_template("showSchedule.html")
    rendered_template = template.render(date=request.args.get('date'), schedule_data=output)
    return html(rendered_template)


@app.post('/admin/DeleteChange')
async def adminPanelDeleteChangePost(request):
    Database.DeleteChange(request.form.get('id'))
    return response.redirect('/admin/ChangesPage?date='+request.args.get('date'))

@app.post('/admin/AddChange')
async def adminPanelAddChangePost(request):
    if request.form.get('buttonValue') == 'edit':
        Database.DeleteChange(int(request.form.get('changeIdinput')))

    date = request.args.get('date')
    start_select = request.form.get('startSelect')
    end_select = request.form.get('endSelect')
    start_change_select = int(request.form.get('startChangeSelect'))
    end_change_select = int(request.form.get('endChangeSelect'))

    for i in range(start_change_select, end_change_select + 1):
        subject = request.form.get('subjectChange')
        prepod = request.form.get('prepodChange')
        group = request.form.get('group')
        date_obj = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        cabinet = request.form.get('cabinetChange')
        change_status_select = request.form.get('changeStatusSelect')
        index = int(start_select) + range(start_change_select, end_change_select + 1).index(i) if start_select else None
        Database.AddChange(subject, prepod, group, date_obj, i, cabinet, change_status_select, index)
    
    return response.redirect('/admin/ChangesPage?date='+date)

def run_another_file():
    import subprocess
    
    # Указать путь к файлу Python, который вы хотите запустить
    python_file_path = "rechecker.py"
    while datetime.datetime.now().strftime("%H:%M") != "04:00":
        pass
    # Запустить другой файл Python
    subprocess.run(["python", python_file_path], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

if __name__ == "__main__":
    print("Запуск")
    thread = threading.Thread(target=run_another_file)
    thread.start()
    app.run(host="0.0.0.0", port=8000)
